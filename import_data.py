import openpyxl
import json
import urllib.request
import urllib.error
from datetime import datetime

FIREBASE_URL = "https://booze-bags-new-default-rtdb.firebaseio.com/orders"

def map_status(excel_status):
    if not excel_status:
        return 'open'
    s = str(excel_status).strip()
    if s in ['אצל הלקוח', 'נשלח', 'הסתיים']:
        return 'closed'
    elif s in ['שלח תמונות ומשפטים', 'מוכן לדפוס', 'הודפס', 'בטיפול']:
        return 'in-progress'
    else:
        return 'open'

def format_date(val):
    if not val:
        return ''
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    return str(val)[:10]

wb = openpyxl.load_workbook('/Users/eggxyt/Downloads/מעקב לקוחות בוזבאגס.xlsx')
ws = wb['לקוחות_ראשי']

orders = []
for row in range(2, ws.max_row + 1):
    name = ws.cell(row, 1).value
    if not name:
        continue

    date_val = ws.cell(row, 2).value
    bags_val = ws.cell(row, 3).value
    phone_val = ws.cell(row, 4).value
    price_val = ws.cell(row, 5).value
    advance_val = ws.cell(row, 6).value
    advance_paid = ws.cell(row, 7).value
    balance_paid = ws.cell(row, 8).value
    status_val = ws.cell(row, 9).value
    cocktails_val = ws.cell(row, 10).value
    location_val = ws.cell(row, 11).value
    delivery_val = ws.cell(row, 12).value
    notes_val = ws.cell(row, 14).value

    # Build bag count
    try:
        bag_count = int(float(bags_val)) if bags_val and bags_val != 'אחר' else 0
    except:
        bag_count = 0

    # Build notes (combine all relevant fields)
    notes_parts = []
    if notes_val:
        notes_parts.append(str(notes_val))
    if cocktails_val:
        notes_parts.append(f'קוקטיילים: {cocktails_val}')
    if location_val:
        notes_parts.append(f'מיקום: {location_val}')
    if delivery_val:
        notes_parts.append(f'הגשה: {delivery_val}')
    if advance_val and advance_val != 0:
        paid_status = f'מקדמה ששולמה: {advance_val}₪'
        if advance_paid:
            paid_status += f' ({advance_paid})'
        if balance_paid:
            paid_status += f' | יתרה: {balance_paid}'
        notes_parts.append(paid_status)

    order = {
        'customerName': str(name).strip(),
        'phone': str(phone_val).strip() if phone_val else '',
        'eventDate': format_date(date_val),
        'eventType': 'חתונה',
        'bagCount': bag_count,
        'packagePrice': float(price_val) if price_val else 0,
        'extras': 0,
        'productionCost': 0,
        'status': map_status(status_val),
        'notes': ' | '.join(notes_parts),
        'createdAt': datetime.now().isoformat()
    }
    orders.append(order)

print(f"נמצאו {len(orders)} הזמנות. מעלה לFirebase...")

success = 0
failed = 0
for i, order in enumerate(orders):
    data = json.dumps(order).encode('utf-8')
    req = urllib.request.Request(
        f"{FIREBASE_URL}.json",
        data=data,
        headers={'Content-Type': 'application/json'},
        method='POST'
    )
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            result = json.loads(resp.read())
            print(f"  ✅ {i+1}/{len(orders)}: {order['customerName']} ({order['eventDate']})")
            success += 1
    except urllib.error.HTTPError as e:
        print(f"  ❌ {i+1}/{len(orders)}: {order['customerName']} - שגיאה: {e.code} {e.reason}")
        failed += 1
    except Exception as e:
        print(f"  ❌ {i+1}/{len(orders)}: {order['customerName']} - שגיאה: {e}")
        failed += 1

print(f"\n✅ הועלו בהצלחה: {success}")
print(f"❌ נכשלו: {failed}")
print(f"\n🎉 רענן את האתר - כל ההזמנות יופיעו!")
